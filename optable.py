from openpyxl import Workbook
from more_itertools import powerset

def tupleIntersect(s, t):
    intrsct = []
    for i in s:
        for j in t:
            if i == j:
                intrsct.append(i)

    return (tuple(intrsct))

def tupleUnion(s, t):
    unn = []
    for i in s:
        for j in t:
            if i == j:
                unn.append(i)
            else:
                unn.append(i)
                unn.append(j)

    return (tuple(unn))


# print (tupleIntersect(('a', (), 'c'), ((), 'c', 'h')))


def belongsTo(tpl, n):
    for i in tpl:
        if n == i:
            return True
        
    return False


def createPowerPowerSetTuple(s):
    oneLift = list(powerset(s))
    oneLift.remove(())
    secLift = list(powerset(oneLift))
    secLift.remove(())
    return secLift

# print (createPowerPowerSetTuple(('a', 'b')))

def createIntersectionTable(postuple, filename):
    wb = Workbook()
    ws = wb.active
    labelList = createPowerPowerSetTuple(postuple)
    for i in range(len(labelList)):
        ws.cell(row = 1, column = i + 2, value = str(labelList[i]))
        ws.cell(row = i + 2, column = 1, value = str(labelList[i]))

    for r in range(len(labelList)):
        for c in range(len(labelList)):
            ws.cell(row = r + 2, column = c + 2, value = str(tupleIntersect(labelList[r], labelList[c])))

    wb.save(filename)

def createUnionTable(postuple, filename):
    wb = Workbook()
    ws = wb.active
    labelList = createPowerPowerSetTuple(postuple)
    for i in range(len(labelList)):
        ws.cell(row = 1, column = i + 2, value = str(labelList[i]))
        ws.cell(row = i + 2, column = 1, value = str(labelList[i]))

    for r in range(len(labelList)):
        for c in range(len(labelList)):
            ws.cell(row = r + 2, column = c + 2, value = str(tupleUnion(labelList[r], labelList[c])))

    wb.save(filename)

    # for row in labelList:
    #     for col in labelList:
    #         ws.cell



# wb = Workbook()
# ws = wb.active
# for i in range(len((1, 2, 3, 4, 5))):
#     ws.cell(row = 1, column = i + 1, value = i + 1)

# wb.save('test.xlsx')
    
createIntersectionTable(('a', 'b'), 'intersect.xlsx')
createUnionTable(('a', 'b'), 'unionTable.xlsx')




# print (list(powerset(('a', 'b'))))
# print ("UP")
# print (list(powerset(list(powerset(('a', 'b'))))))